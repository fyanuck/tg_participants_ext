import json
import os
import datetime
from telethon.sync import TelegramClient
from telethon.tl.functions.channels import GetParticipantsRequest
from telethon.tl.types import ChannelParticipantsSearch
import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


def get_client():
    # api_id, api_hash из JSON
    json_fname = 'API_AUTH.json'
    if not os.path.exists(json_fname):
        print(f'Ошибка чтения JSON: нет файла с именем {json_fname}')
        exit(1)

    json_dict = {}
    with open(json_fname, 'r') as f:
        try:
            json_dict = json.load(f)
        except json.JSONDecodeError as e:
            print(f'Ошибка чтения JSON: {e}')
            exit(1)

    if any(field not in json_dict.keys() for field in ('api_id', 'api_hash')):
        print('Ошибка чтения JSON: необходимы поля api_id, api_hash')
        exit(1)

    api_id = json_dict['api_id']
    api_hash = json_dict['api_hash']

    client = TelegramClient('session_name', api_id, api_hash)
    client.start()
    return client


def get_participants(client, entity_id: str):
    try:
        people = []   # Список словарей вида {id, username, first_name, last_name}
        offset = 0

        entity = client.get_entity(entity_id)
        while True:
            participants = client(GetParticipantsRequest(
                entity, ChannelParticipantsSearch(''), offset, limit=200, hash=0))
            if not participants.users:
                break
            for person in participants.users:
                people.append({'id': person.id, 'username': person.username, 
                              'first_name': person.first_name, 'last_name': person.last_name})
            offset += len(participants.users)
        return people
    except Exception as e:
        print(f'Ошибка списка участников: {e}')
        exit(1)


def save_people(people, entity_id):
    wb = openpyxl.Workbook()
    ws = wb.active
    now_str = datetime.datetime.now().strftime('%d.%m.%Y %H:%M:%S')
    now_date_str = now_str.split()[0]
    entity_str = entity_id.split('/')[-1]
    ws.title = f'{entity_str} {now_date_str}'

    ws.cell(row=1, column=1).value = f'Список участников Telegram {entity_str} на {now_str}'
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
    ws.cell(row=2, column=1).value = '#'
    ws.cell(row=2, column=2).value = 'id'
    ws.cell(row=2, column=3).value = 'username'
    ws.cell(row=2, column=4).value = 'Имя'
    ws.cell(row=2, column=5).value = 'Фамилия (если есть)'

    for r in (1, 2):
        for c in (1, 2, 3, 4, 5):
            ws.cell(row=r, column=c).font = Font(bold=True)

    for i, person in enumerate(people, start=1):
        ws.cell(row=2+i, column=1).value = i
        ws.cell(row=2+i, column=2).value = person['id']
        ws.cell(row=2+i, column=3).value = person['username']
        ws.cell(row=2+i, column=4).value = person['first_name']
        if person['last_name'] is not None:
            ws.cell(row=2+i, column=5).value = person['last_name']

    for idx, col in enumerate(ws.columns, 1):
        ws.column_dimensions[get_column_letter(idx)].auto_size = True
    
    savename = f'Участники {entity_str}.xlsx'
    copy_counter = 1
    while os.path.exists(savename):
        savename = f'Участники {entity_str} ({copy_counter}).xlsx'
        copy_counter += 1
    wb.save(savename)
    return savename


def main():
    with open('ID.txt', 'r') as f:
        entity_id = f.readline().strip()
    
    client = get_client()
    people = get_participants(client, entity_id)
    savename = save_people(people, entity_id)
    print(f'Готово! Сохранено в {savename}')


if __name__ == '__main__':
    main()